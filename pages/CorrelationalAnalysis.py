import dash_core_components as dcc
import dash_html_components as html
import plotly.express as px
import dash_table


import dash
from dash.dependencies import Input, Output
from app import app, data, INDEX_YEAR, MIN_YEAR

#from utils import Header
from utils import Header, Footer

import numpy as np
import pandas as pd
# new imports
import pages.corr_analysis_tool as cat
import plots.network_analysis.igraph_to_plotly as itop
import data_utils as du
from utils import is_btn_clicked
from ggmodel_dev.models.greengrowth import GGGM
import plotly.graph_objects as go
import plotly.express as px
from sklearn.manifold import trustworthiness
from openpyxl import load_workbook

networks = GGGM.all_model_dictionary

#%%
corr_types = []
corr_types.append({'label': 'Direct', 'value': 'Direct'})
#corr_types.append({'label': 'Indirect', 'value': 'Indirect'})
#corr_types.append({'label': 'Causal-Correlational', 'value': 'Causal-Correlational'})

wb = load_workbook("./data/causal_data.xlsx", read_only=True, keep_links=False)
imp_models = wb.sheetnames

network_options = []

for i in imp_models:
    network_options.append({"label" : i, "value" : i})

#data = pd.read_excel("./data/causal_data.xlsx", engine='openpyxl', index_col =0, sheet_name ="FPi_model")
# var_types = data.loc[:, 'Type']
# data.drop(labels = 'Type', axis = 1, inplace=True)
# rem_vars = data[data.isna().all(axis=1)].index
# data.drop(rem_vars, axis = 0, inplace=True)
# data.dropna(inplace=True, axis = 1)
# X = data.transpose()
# X = (X - X.min(axis=0)) / (X.max(axis=0) - X.min(axis=0))

#%%
layout = html.Div(
    [
        html.Div([
            Header(app, 'Correlational Analysis')
        ]),
        html.Div(
            [
                html.Div(
                    [
                       html.Div(
                            [
                                html.H5("Highlights"),
                                html.Br([]),
                                html.P("Correlational analysis measures the relationship between input and output variables. Generally speaking, the higher the correlation is, the more similarly change two variables. However, it is important to note that correlation is not causality.",
                                       style={"color": "#ffffff", 'font-size': '15px'},),
                                html.Br([]),
                                html.P("This page vizualizes the correlations as direct networks. The direct correlation measures the correlation between the input and output variables.", #and indirect graphs , whereas the indirect measures the correlation through the paths of the network
                                       style={"color": "#ffffff", 'font-size': '15px'},)
                                       
                            ],
                            className="product"
                        ),
                        html.Div(
                            [
                                html.Div(children='Select Correlation Mode'),
                                dcc.Dropdown(id='ca_corr_sel_dropdown', options=corr_types, value = corr_types[0]['value']),
                                html.Br([]),
                                dcc.Dropdown(id='ca_network_sel_dropdown', options=network_options, value = network_options[0]['value']),
                                html.Br([]),
                                html.Div(id="placeholder"),
                                #dcc.Dropdown(id='inp_network_sel_dropdown', options=corr_types, value = corr_types[0]['value']),
                                html.Button('Run', id='btn-run-ca', n_clicks=0,
                                style={'font-size': 20,
                                       'font-weight': 'normal',
                                       'color': '#ffffff',
                                       'background': '#14ac9c',
                                       'border': '#14ac9c',
                                       }) #,
                                #html.Br([]),
                                #html.Button('Download', id='btn-dwn-qa', n_clicks=0,
                                #style={'font-size': 20,
                                #       'font-weight': 'normal',
                                #       'color': '#ffffff',
                                #       'background': '#14ac9c',
                                #       'border': '#14ac9c',
                                #       }),
                                #dcc.Download(id="download-qa")

                            ], className="na_ui"

                        )
                    ],
                    className="pretty_container four columns"
                ),
                html.Div(
                    [
                        html.H5("Correlational Analysis results:"),
                        html.Br([]),
                        html.Div(id = "c_results_id", children = dcc.Graph(id='corr_fig',figure=go.Figure(), responsive=True))
                    ],
                    className="pretty_container eight columns"
                ),

            ],
            className="row",
        ),
        Footer(),
    ],
    className="page",
)
# @app.callback(
#     Output('outp_dropdown', 'options'),
#     Output('outp_dropdown', 'value'),
#     Input('qa_network_sel_dropdown', 'value')
# )
# def SetNodes(input):
#     sel_vars = []
#     vs = list(data[input].keys())
#     for y in vs:
#         sel_vars.append({'label': data[input][y]['name'], 'value': y})

#     sel_var_lab = None

#     if len(sel_vars) != 0:
#         sel_var_lab = sel_vars[0]["value"]

#     return sel_vars, sel_var_lab

@app.callback(
    Output('corr_fig', 'figure'),
    Input('ca_corr_sel_dropdown', 'value'),
    Input('ca_network_sel_dropdown', 'value')
) 
def DirectCorrelationVizualization(corr_typ, model):

    data = pd.read_excel("./data/causal_data.xlsx", engine='openpyxl', index_col =0, sheet_name =model)

    rem_vars = data[data.isna().all(axis=1)].index.tolist()
    data.drop(rem_vars, axis = 0, inplace=True)
    data.dropna(inplace=True, axis = 1)

    var_types = data.loc[:, 'Type']
    data.drop(labels = 'Type', axis = 1, inplace=True)
        
    X = data.transpose()
    X = (X - X.min(axis=0)) / (X.max(axis=0) - X.min(axis=0))

    figa =go.Figure()
    if corr_typ == 'Direct':
        G, cl, labels, ann_text  = cat.DirectCorrelation(X, rem_vars, var_types)
        figa = itop.Vizualize_iGraph_Plotly(G, labels, cl, annotation_text=ann_text, marker_annonation=False)
    return figa